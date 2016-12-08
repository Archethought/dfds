import sys

"""
Show how to read Excel spreadsheet

Program takes a command line parameter

Usage: python readxls.py some_excel.xlsx

N.B. There is a bug in openpyxl that looks like this:

---

Traceback (most recent call last):
  File "readxls.py", line 16, in <module>
    wb = openpyxl.load_workbook(sys.argv[1])
  File "/Users/dixon/working/dfds/sheets/venv/lib/python2.7/site-packages/openpyxl/reader/excel.py", line 161, in load_workbook
    parser.parse()
  File "/Users/dixon/working/dfds/sheets/venv/lib/python2.7/site-packages/openpyxl/packaging/workbook.py", line 43, in parse
    if package.properties.date1904:
AttributeError: 'NoneType' object has no attribute 'date1904'

---

To get around this requires checking out version 2.4 from bitbucket.com repo
https://bitbucket.org/openpyxl/openpyxl/src/509eb2a6775b1b126794f36cf2ecce86daf1d9e9/?at=2.4

Add the directory 'openpyxl' to your code directory and then import will find it using the sys.path.append

FYI how to checkout on bitbucket.com:
https://confluence.atlassian.com/bitbucket/checkout-a-branch-into-a-local-repository-313466957.html

"""

sys.path.append('./openpyxl')
# print sys.path

import openpyxl
from openpyxl import Workbook

### Create a workbook
wb = Workbook()

wb = openpyxl.load_workbook(sys.argv[1])

### Get the active worksheets (index 0)
ws = wb.active

### get the dimensions and print them
mr = ws.max_row
mc = ws.max_column

print mr, mc

### Print the data for the whole sheet
for row in ws.iter_rows():
 for cell in row:
   if cell.value is not None:
     print cell.value,
 print
